"""
Module for helper functions
"""
import os.path
import sys
import time
import datetime
import calendar
# from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Border, Side

import pandas as pd
import json

import logging
logger = logging.getLogger(__name__)


class helper:
    """helper class"""

    def __init__(self) -> None:
        pass

    def file_name(self, name_initial):
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        name = name_initial + timestamp + '.xlsx'
        return name

    def file_path(self, name_initial="Sales_Digest_"):
        temp_dir = './temp'
        if not os.path.exists(temp_dir):
            os.makedirs(temp_dir, exist_ok=True)
        name = self.file_name(name_initial)
        path = os.path.join(temp_dir, name)
        return path, name

    def convert_utc_date(self, date_str):
        # To accommodate various date pattern
        date_str=date_str.split("T")[0]
        # dateformat = '%Y-%m-%dT%H:%M:%S.%fZ'
        dateformat = '%Y-%m-%d'
        date_obj = datetime.datetime.strptime(date_str, dateformat)
        formatted_date = date_obj.strftime("%A, %B %d, %Y")
        return formatted_date

    def create_dataframe(self, data):
        df = pd.DataFrame(data)
        df['news_date'] = df['news_date'].apply(self.convert_utc_date)
        # df.loc[df['source_name'] == 'Drug Channel', 'news_date'] = df.loc[
        #     df['source_name'] == 'Drug Channel', 'news_date'].apply(self.convert_utc_date)

        #Create Capital for Sentiment
        df['sentiment']=df['sentiment'].apply(lambda x: x.capitalize())

        df = df.rename(columns={'news_date': 'Date', "news_title": "Title", "news_url": "URL",
                                "news_summary": "Summary",
                                "client_name": "Client",
                                "source_name": "Source",
                                "sentiment": "Sentiment"})
        df.index += 1
        return df

    def create_excel(self, df, f_path):
        workbook = Workbook()
        sheet = workbook.active

        # Set headers in the first row
        headers = ['Sr NO.'] + list(df.columns)
        sheet.append(headers)

        # Set dynamic column widths and apply text wrapping
        max_width = 78
        wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for col_idx, col in enumerate(headers, start=1):
            max_length = max(df[col].astype(str).apply(len).max(), len(col)) if col not in ['Sr NO.',
                                                                                            'Sentiment'] else len(
                col) + 1
            adjusted_width = min(max_length, max_width)
            column_letter = chr(64 + col_idx)
            sheet.column_dimensions[column_letter].width = adjusted_width

        # Add data from DataFrame to worksheet and apply text wrapping
        for row in dataframe_to_rows(df.reset_index(), index=False, header=False):
            sheet.append(row)

        # Apply text wrapping to all cells
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = wrap_alignment
                cell.border = thin_border

        # Save workbook to file
        workbook.save(f_path)

    def delete_file_after_delay(self, f_path):
        try:
            delay = os.getenv('DELETE_FILE_DELAY')
            time.sleep(int(delay))
            if os.path.exists(f_path):
                os.remove(f_path)
                logger.info(f"File deleted after {delay} seconds: {f_path}")
            else:
                logger.warning("File not found: " + f_path)
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            logger.error(f"Error Deleting file: {f_path}Line No: {exc_tb.tb_lineno}  Error: {str(e)}", stack_info=True)

    def start_end_date(self, date):
        """
        function to return start date and end date from the input date which is of format YYYY-MM
        """
        try:
            dateformat = '%Y-%m-%dT%H:%M:%S.%fZ'
            # Split the date into year and month
            year, month = map(int, date.split('-'))

            # Start date is always the first day of the month
            start_date = datetime.datetime(year, month, 1)

            # Use calendar.monthrange() to get the number of days in the month
            _, num_days = calendar.monthrange(year, month)

            # End date is the last day of the month
            end_date = datetime.datetime(year, month, num_days)
            start_date = start_date.strftime(dateformat)
            end_date = end_date.strftime(dateformat)

            return start_date, end_date
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            logger.error(f"Error in converting dates. Line No: {exc_tb.tb_lineno}  Error: {str(e)}", stack_info=True)

    def deduplicate_dicts(self, dict_list, keyword):
        seen = set()
        deduplicated_list = [d for d in dict_list if not (d[keyword] in seen or seen.add(d[keyword]))]
        return deduplicated_list

    def create_definitive_excel(self, data, f_path):
        with pd.ExcelWriter(f_path, engine='openpyxl') as writer:
            for category, metrics in data.items():
                # Create a dataframe for each category
                df = pd.DataFrame(list(metrics.items()), columns=["Metrics", "Performance"])

                #Add Serial no. column
                df.insert(0, "S.No.", range(1, len(df) + 1))
                # Write the dataframe to a different sheet
                df.to_excel(writer, sheet_name=category, index=False)

        # Load the workbook to adjust column widths
        workbook = load_workbook(f_path)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            # Calculate the maximum width needed for the "Parameter" and "Value" columns
            max_param_width = max([len(str(cell.value)) for cell in sheet['B']] + [len("Parameter")])
            max_value_width = max([len(str(cell.value)) for cell in sheet['C']] + [len("Value")])
            # Set the width of the "Parameter" and "Value" columns
            sheet.column_dimensions['A'].width = 6
            sheet.column_dimensions['B'].width = max_param_width + 1
            sheet.column_dimensions['C'].width = max_value_width + 1
            wrap_alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            # Create a thin border around the cells
            for row in sheet.iter_rows():
                for cell in row:
                    border = Border(left=Side(style='thin'),
                                    right=Side(style='thin'),
                                    top=Side(style='thin'),
                                    bottom=Side(style='thin'))
                    cell.border = border
                    cell.alignment = wrap_alignment
            # Save the workbook with adjusted column widths
        workbook.save(f_path)
        
    def add_feedback_indicator(self,news_data,feedbacks):
        try:
            source_data_df = pd.DataFrame(news_data)
            feedbacks_df = pd.DataFrame(feedbacks)
            result_df = pd.merge(source_data_df, feedbacks_df[['news_url','feedback']], on=['news_url'], how='left')
            #result_df[['isThumbsUp','isThumbsDown']] = False
            result_df.loc[result_df['feedback']=='positive','isThumbsUp'] = True
            result_df.loc[result_df['feedback']=='negative','isThumbsDown'] = True
            result_df.drop(columns=['feedback'],axis=1,inplace=True)
            result = json.loads(json.dumps(list(result_df.T.to_dict().values())))
 
           
            return result
        except Exception as e:
            exc_type, exc_obj, exc_tb = sys.exc_info()
            logger.error(f"Error while adding feedback indicator:   Error: {str(e)}", stack_info=True)